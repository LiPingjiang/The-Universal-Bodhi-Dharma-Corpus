#!/usr/bin/env python3
import openpyxl, os, glob

wb = openpyxl.load_workbook('佛教佛法文献古籍大全.xlsx')
ws = wb['文献大全']

excel_names = []
for r in range(2, ws.max_row + 1):
    num = ws.cell(r, 1).value
    period = ws.cell(r, 2).value
    era = ws.cell(r, 3).value
    name = ws.cell(r, 4).value
    excel_names.append((num, period, era, name))

md_files = glob.glob('文献档案/*.md')
md_names = [os.path.basename(f).replace('.md', '') for f in md_files if 'README' not in f]

print('=== Excel条目 vs 文件名匹配分析 ===\n')

mapping = {}
for num, period, era, name in excel_names:
    found = None
    for md in md_names:
        parts = md.split('_', 1)
        md_name = parts[1] if len(parts) > 1 else md
        md_simple = md_name.replace('（', '').replace('）', '').replace('与', '').replace('及', '')
        name_simple = str(name).replace('（', '').replace('）', '').replace('与', '').replace('及', '') if name else ''
        if name_simple and md_simple:
            if name_simple in md_simple or md_simple in name_simple or name_simple == md_simple:
                found = md
                break
    mapping[num] = (name, found)

matched = sum(1 for v in mapping.values() if v[1] is not None)
unmatched = sum(1 for v in mapping.values() if v[1] is None)

print(f'Excel总条目: {len(excel_names)}')
print(f'已匹配: {matched}')
print(f'未匹配: {unmatched}')
print(f'文件名总数: {len(md_names)}')
print()

print('=== 未匹配的Excel条目 ===')
for num, (name, found) in mapping.items():
    if found is None:
        print(f'{num}: {name}')

print('\n=== 未匹配的文件名（可能多余）===')
matched_mds = set(v[1] for v in mapping.values() if v[1] is not None)
for md in md_names:
    if md not in matched_mds:
        print(md)
